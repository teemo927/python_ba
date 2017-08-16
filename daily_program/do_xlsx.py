#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import xlsxwriter


# 使用
class DoXlsx(object):
    def __init__(self, xls_name):
        self.xls_name = xls_name

    def pre(self, txt):
        wb = load_workbook(txt)
        print(wb.sheetnames)
        sht = wb.get_active_sheet()
        print(sht)

    # 新建xlsx文件
    def mk_xls(self, file_name, sheet_name='Sheet1'):
        import os
        if not os.path.exists(os.getcwd() + '/' + file_name):
            workbook = xlsxwriter.Workbook(file_name)
            worksheet = workbook.add_worksheet(sheet_name)
            worksheet2 = workbook.add_worksheet('娃哈哈')
            # expenses = (
            #     ['Rent', 1000],
            #     ['Gas', 100],
            #     ['Food', 300],
            #     ['Gym', 50],
            # )
            expenses = (
                ["张三", 150, 120, 100],
                ["李四", 90, 99, 95],
                ["王五", 60, 66, 68],
            )
            row = 0
            col = 0
            # Iterate over the data and write it out row by row.
            for item, cost, cost2, cost3 in (expenses):
                worksheet.write(row, col, item)
                worksheet.write(row, col + 1, cost)
                worksheet.write(row, col + 2, cost2)
                worksheet.write(row, col + 3, cost3)
                row += 1

            # Write a total using a formula.
            worksheet.write(row, 0, 'Total')
            worksheet.write(row, 1, '=SUM(B1:B3)')  # 调用excel的公式表达式
            worksheet.write(row, 2, '=SUM(C1:C3)')  # 调用excel的公式表达式
            worksheet.write(row, 3, '=SUM(D1:D3)')  # 调用excel的公式表达式

            worksheet.write(0, 4, '=SUM(B1:D1)')  # 调用excel的公式表达式
            worksheet.write(1, 4, '=SUM(B2:D2)')  # 调用excel的公式表达式
            worksheet.write(2, 4, '=SUM(B3:D3)')  # 调用excel的公式表达式
            worksheet.write(3, 4, '=SUM(B4:D3)')  # 调用excel的公式表达式
            workbook.close()
        else:
            print('file is exists: ', file_name)


if __name__ == '__main__':
    name = 'temp.xlsx'
    do = DoXlsx(name)
    do.mk_xls(name, '薇薇安')

    do.pre(name)
